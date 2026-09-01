import os
import io
import re
import tempfile
import subprocess
import openpyxl
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.graphic import GroupShape

PLANTILLA_EXCEL = "PLANTILLA_GR.xlsx"

COPIAS_POR_AREA = {
    "SUPERVISION": 2,
    "CALIDAD": 3,
    "TOPOGRAFIA": 2,
    "PRODUCCION": 2
}

def obtener_texto_celda_b6(area):
    area_norm = area.upper().strip()
    if area_norm == "SUPERVISION":
        return "SUPERVISION"
    return f"{area_norm} OSP"

def extraer_numero_revision(codigo_plano):
    match = re.search(r'-(?:R|REV)(\d+)$', codigo_plano, re.IGNORECASE)
    if match:
        return int(match.group(1))
    match_alt = re.search(r'-(\d+)$', codigo_plano)
    if match_alt:
        return int(match_alt.group(1))
    return ""

def calcular_siguiente_correlativo(secuencia_base, offset):
    match = re.match(r'^(\d+)(-.+)$', secuencia_base.strip())
    if match:
        num_actual = int(match.group(1))
        resto = match.group(2)
        nuevo_num = str(num_actual + offset)
        return nuevo_num, resto
    return secuencia_base, ""

def agregar_linea_diagonal_excel(ws, fila_inicio_linea, fila_fin_linea=44):
    if fila_inicio_linea >= fila_fin_linea:
        return
    try:
        marker_from = AnchorMarker(col=1, colOff=0, row=fila_inicio_linea - 1, rowOff=0)
        marker_to = AnchorMarker(col=10, colOff=0, row=fila_fin_linea, rowOff=0)
        anchor = TwoCellAnchor(_from=marker_from, to=marker_to)
        shape = GroupShape()
        anchor.sp = shape
        ws.add_drawing(anchor)
    except Exception:
        pass

def generar_excel_por_area(resumen_planos, fecha_texto, secuencia_base, area, offset_correlativo, plantilla_path=PLANTILLA_EXCEL):
    if not os.path.exists(plantilla_path):
        return None, None

    wb = openpyxl.load_workbook(plantilla_path)

    if "osp" in wb.sheetnames:
        ws = wb["osp"]
        wb.active = ws
        for sheet in wb.sheetnames:
            if sheet != "osp":
                del wb[sheet]
    else:
        ws = wb.active

    ws.cell(row=6, column=2, value=obtener_texto_celda_b6(area))

    num_str, resto_str = calcular_siguiente_correlativo(secuencia_base, offset_correlativo)
    secuencia_completa = f"{num_str}{resto_str}"

    font_bold = InlineFont(b=True, rFont="Calibri", sz=11)
    font_normal = InlineFont(b=False, rFont="Calibri", sz=11)

    rich_i5 = CellRichText([
        TextBlock(font_bold, num_str),
        TextBlock(font_normal, resto_str)
    ])
    
    ws.cell(row=5, column=9, value=rich_i5)
    ws.cell(row=5, column=10, value=fecha_texto)

    fila_inicio = 10
    num_copias = COPIAS_POR_AREA.get(area.upper(), 2)

    for idx, plano in enumerate(resumen_planos):
        fila = fila_inicio + idx
        cod = plano["Código de Plano"]
        desc = plano["Título / Descripción"]
        rev_num = extraer_numero_revision(cod)

        ws.cell(row=fila, column=2, value=cod)
        ws.cell(row=fila, column=4, value=desc)
        ws.cell(row=fila, column=8, value=rev_num)
        ws.cell(row=fila, column=9, value=num_copias)
        ws.cell(row=fila, column=10, value=5)

    ultima_fila_usada = fila_inicio + len(resumen_planos) - 1
    agregar_linea_diagonal_excel(ws, ultima_fila_usada + 1, fila_fin_linea=44)

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream.getvalue(), secuencia_completa

def convertir_excel_a_pdf(excel_bytes):
    tmp_excel_path = None
    expected_pdf_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            tmp_excel.write(excel_bytes)
            tmp_excel_path = tmp_excel.name

        out_dir = tempfile.gettempdir()
        cmd = ["libreoffice", "--headless", "--convert-to", "pdf", tmp_excel_path, "--outdir", out_dir]
        subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)

        expected_pdf_path = os.path.splitext(tmp_excel_path)[0] + ".pdf"

        if os.path.exists(expected_pdf_path):
            with open(expected_pdf_path, "rb") as f:
                pdf_bytes = f.read()
            os.remove(tmp_excel_path)
            os.remove(expected_pdf_path)
            return pdf_bytes, None
        return None, "No se generó el PDF con LibreOffice."
    except Exception as e:
        if tmp_excel_path and os.path.exists(tmp_excel_path):
            os.remove(tmp_excel_path)
        if expected_pdf_path and os.path.exists(expected_pdf_path):
            os.remove(expected_pdf_path)
        return None, str(e)