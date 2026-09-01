import streamlit as st
import fitz       # PyMuPDF
import cv2
import numpy as np
import tempfile
import os
import glob
import re
import datetime
import io
import openpyxl
import subprocess
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.graphic import GroupShape

# --- CONFIGURACIÓN DE MEDIDAS EXACTAS EN PAPEL (CENTÍMETROS REALES) ---
CM_TO_PT = 28.34645669

ANCHO_SELLO_CM = 7.1
ALTO_SELLO_CM = 2.6

CARPETA_SELLOS = "firmas_sellos"
PLANTILLA_EXCEL = "PLANTILLA_GR.xlsx"
RUTA_LOGO = "logo.png"

if not os.path.exists(CARPETA_SELLOS):
    os.makedirs(CARPETA_SELLOS)

# --- INYECCIÓN DE ESTILOS CSS PERSONALIZADOS (TEMA OSCURO CORPORATIVO) ---
# --- INYECCIÓN DE ESTILOS CSS PERSONALIZADOS (TEMA OSCURO CORPORATIVO) ---
def aplicar_estilos_custom():
    st.markdown("""
        <style>
        /* Fondo general y color de texto base */
        .stApp {
            background-color: #0B132B;
            color: #F8FAFC;
        }

        /* Barra Lateral (Sidebar) */
        [data-testid="stSidebar"] {
            background-color: #1C2541;
            border-right: 1px solid #1E293B;
        }
        
        [data-testid="stSidebar"] * {
            color: #F8FAFC !important;
        }

        /* Estilo para los títulos y etiquetas */
        h1, h2, h3, h4, label, .stMarkdown p {
            color: #F8FAFC !important;
            font-weight: 600;
        }
        
        .stCaption {
            color: #94A3B8 !important;
        }

        /* Contenedor principal */
        .main .block-container {
            padding-top: 2rem;
            max-width: 95%;
        }

        /* Entradas de texto, fechas y selectores */
        .stTextInput input, .stDateInput input, [data-baseweb="select"], [data-baseweb="base-input"] {
            background-color: #1C2541 !important;
            color: #F8FAFC !important;
            border: 1px solid #3B82F6 !important;
            border-radius: 6px !important;
        }

        /* FIX: MultiSelect tags (Etiquetas seleccionadas en Azul Corporativo) */
        span[data-baseweb="tag"], 
        div[data-baseweb="tag"],
        [data-baseweb="tag"] {
            background-color: #005596 !important;
            border-color: #0077C8 !important;
            border-radius: 4px !important;
        }
        
        span[data-baseweb="tag"] *, 
        [data-baseweb="tag"] span {
            color: #FFFFFF !important;
        }

        /* Checkbox color acento azul */
        input[type="checkbox"]:checked {
            background-color: #005596 !important;
            border-color: #005596 !important;
        }

        /* Uploader de archivos */
        [data-testid="stFileUploader"] {
            background-color: #1C2541;
            border: 1px dashed #38BDF8;
            border-radius: 8px;
            padding: 10px;
        }

        /* Botones Principales (Azul Corporativo) */
        .stButton>button {
            background: linear-gradient(90deg, #005596, #0077C8);
            color: #FFFFFF !important;
            border: none;
            border-radius: 6px;
            font-weight: bold;
            padding: 0.6rem 1.2rem;
            transition: all 0.3s ease;
        }
        
        .stButton>button:hover {
            background: linear-gradient(90deg, #003D6D, #005596);
            box-shadow: 0 4px 12px rgba(56, 189, 248, 0.3);
            color: #FFFFFF !important;
        }

        /* Botones de Descarga */
        .stDownloadButton>button {
            background-color: #0F172A;
            color: #38BDF8 !important;
            border: 1px solid #38BDF8;
            border-radius: 6px;
            font-weight: bold;
            transition: all 0.3s ease;
        }
        
        .stDownloadButton>button:hover {
            background-color: #005596;
            color: #FFFFFF !important;
            border: 1px solid #005596;
        }

        /* Estilo para las Tablas de datos */
        [data-testid="stDataFrame"] {
            background-color: #1C2541;
            border-radius: 8px;
            border: 1px solid #334155;
        }

        /* Formulario de Login */
        [data-testid="stForm"] {
            background-color: #1C2541;
            border: 1px solid #334155;
            border-radius: 12px;
            padding: 20px;
        }

        /* Divisores horizontales */
        hr {
            border-color: #334155 !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
# --- MÓDULO DE AUTENTICACIÓN ---
def inicializar_estado_sesion():
    if "autenticado" not in st.session_state:
        st.session_state.autenticado = False
    if "usuario" not in st.session_state:
        st.session_state.usuario = None

def validar_credenciales(usuario, clave):
    if usuario == "admin" and clave == "admin":
        return True
    return False

def vista_login():
    aplicar_estilos_custom()
    
    col_a, col_b, col_c = st.columns([1, 1.8, 1])
    with col_b:
        if os.path.exists(RUTA_LOGO):
            st.image(RUTA_LOGO, use_container_width=True)
        else:
            st.markdown("<h2 style='text-align: center; color: #38BDF8;'>AO RNLD US</h2>", unsafe_allow_html=True)
            
        st.markdown("<h4 style='text-align: center; color: #94A3B8; font-weight: 300;'>INGENIERÍA • DISEÑO • CONSTRUCCIÓN</h4>", unsafe_allow_html=True)
        st.write("")

        with st.form("login_form"):
            user_input = st.text_input("Usuario", placeholder="Ingrese su usuario")
            pass_input = st.text_input("Contraseña", type="password", placeholder="••••••••")
            submit = st.form_submit_button("INGRESAR AL SISTEMA", use_container_width=True)
            
            if submit:
                if validar_credenciales(user_input, pass_input):
                    st.session_state.autenticado = True
                    st.session_state.usuario = user_input
                    st.success("Acceso Autorizado")
                    st.rerun()
                else:
                    st.error("Credenciales incorrectas")

# --- FUNCIONES DE BASE DE DATOS DE SELLOS Y EXCEL ---
def obtener_libreria_sellos():
    extensiones = ('*.png', '*.jpg', '*.jpeg', '*.PNG', '*.JPG')
    archivos = []
    for ext in extensiones:
        archivos.extend(glob.glob(os.path.join(CARPETA_SELLOS, ext)))

    libreria = {}
    for ruta in archivos:
        nombre_base = os.path.basename(ruta)
        nombre_sin_ext = os.path.splitext(nombre_base)[0].replace("_", " ").title()
        libreria[nombre_sin_ext] = ruta
    return libreria

COPIAS_POR_AREA = {
    "SUPERVISION": 2,
    "CALIDAD": 3,
    "TOPOGRAFIA": 2,
    "PRODUCCION": 2
}

def obtener_texto_celda_b6(area):
    area_norm = area.upper().strip()
    if area_norm == "SUPERVISION":
        return "SUPERVISION"
    return f"{area_norm} OSP"

def extraer_numero_revision(codigo_plano):
    match = re.search(r'-(?:R|REV)(\d+)$', codigo_plano, re.IGNORECASE)
    if match:
        return int(match.group(1))
    match_alt = re.search(r'-(\d+)$', codigo_plano)
    if match_alt:
        return int(match_alt.group(1))
    return ""

def calcular_siguiente_correlativo(secuencia_base, offset):
    match = re.match(r'^(\d+)(-.+)$', secuencia_base.strip())
    if match:
        num_actual = int(match.group(1))
        resto = match.group(2)
        nuevo_num = str(num_actual + offset)
        return nuevo_num, resto
    return secuencia_base, ""

def agregar_linea_diagonal_excel(ws, fila_inicio_linea, fila_fin_linea=44):
    if fila_inicio_linea >= fila_fin_linea:
        return

    try:
        marker_from = AnchorMarker(col=1, colOff=0, row=fila_inicio_linea - 1, rowOff=0)
        marker_to = AnchorMarker(col=10, colOff=0, row=fila_fin_linea, rowOff=0)
        anchor = TwoCellAnchor(_from=marker_from, to=marker_to)
        shape = GroupShape()
        anchor.sp = shape
        ws.add_drawing(anchor)
    except Exception:
        pass

def convertir_excel_a_pdf(excel_bytes):
    tmp_excel_path = None
    expected_pdf_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            tmp_excel.write(excel_bytes)
            tmp_excel_path = tmp_excel.name

        out_dir = tempfile.gettempdir()
        cmd = ["libreoffice", "--headless", "--convert-to", "pdf", tmp_excel_path, "--outdir", out_dir]
        subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)

        expected_pdf_path = os.path.splitext(tmp_excel_path)[0] + ".pdf"

        if os.path.exists(expected_pdf_path):
            with open(expected_pdf_path, "rb") as f:
                pdf_bytes = f.read()
            os.remove(tmp_excel_path)
            os.remove(expected_pdf_path)
            return pdf_bytes, None
        else:
            return None, "No se encontró el archivo PDF generado por LibreOffice."
    except Exception as e:
        if tmp_excel_path and os.path.exists(tmp_excel_path):
            os.remove(tmp_excel_path)
        if expected_pdf_path and os.path.exists(expected_pdf_path):
            os.remove(expected_pdf_path)
        return None, f"Error en conversión: {str(e)}."

def unificar_pdfs(lista_pdf_bytes):
    pdf_final = fitz.open()
    for b in lista_pdf_bytes:
        if b:
            doc_temp = fitz.open(stream=b, filetype="pdf")
            pdf_final.insert_pdf(doc_temp)
            doc_temp.close()

    buffer = io.BytesIO()
    pdf_final.save(buffer)
    pdf_final.close()
    buffer.seek(0)
    return buffer.getvalue()

def generar_excel_por_area(resumen_planos, fecha_texto, secuencia_base, area, offset_correlativo, plantilla_path=PLANTILLA_EXCEL):
    if not os.path.exists(plantilla_path):
        st.error(f"No se encontró la plantilla `{plantilla_path}`.")
        return None, None

    wb = openpyxl.load_workbook(plantilla_path)

    if "osp" in wb.sheetnames:
        ws = wb["osp"]
        wb.active = ws
        for sheet in wb.sheetnames:
            if sheet != "osp":
                del wb[sheet]
    else:
        ws = wb.active

    ws.cell(row=6, column=2, value=obtener_texto_celda_b6(area))

    num_str, resto_str = calcular_siguiente_correlativo(secuencia_base, offset_correlativo)
    secuencia_completa = f"{num_str}{resto_str}"

    font_bold = InlineFont(b=True, rFont="Calibri", sz=11)
    font_normal = InlineFont(b=False, rFont="Calibri", sz=11)

    rich_i5 = CellRichText([
        TextBlock(font_bold, num_str),
        TextBlock(font_normal, resto_str)
    ])
    
    ws.cell(row=5, column=9, value=rich_i5)
    ws.cell(row=5, column=10, value=fecha_texto)

    fila_inicio = 10
    num_copias = COPIAS_POR_AREA.get(area.upper(), 2)
    cant_items = len(resumen_planos)

    for idx, plano in enumerate(resumen_planos):
        fila = fila_inicio + idx
        cod = plano["Código de Plano"]
        desc = plano["Título / Descripción"]
        rev_num = extraer_numero_revision(cod)

        ws.cell(row=fila, column=2, value=cod)
        ws.cell(row=fila, column=4, value=desc)
        ws.cell(row=fila, column=8, value=rev_num)
        ws.cell(row=fila, column=9, value=num_copias)
        ws.cell(row=fila, column=10, value=5)

    ultima_fila_usada = fila_inicio + cant_items - 1
    fila_diagonal_inicio = ultima_fila_usada + 1

    agregar_linea_diagonal_excel(ws, fila_diagonal_inicio, fila_fin_linea=44)

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream.getvalue(), secuencia_completa

# --- EXTRACCIÓN Y BÚSQUEDA ---
def limpiar_texto_comillas(texto):
    if not texto:
        return ""
    texto_limpio = re.sub(r'["“"”\']', '', texto).strip()
    texto_limpio = re.sub(r'^\s*[\-\–\—\:]+\s*', '', texto_limpio)
    return texto_limpio.strip()

def extraer_codigo_plano(pagina):
    mat_view = pagina.rotation_matrix
    view_rect = pagina.rect * mat_view
    v_width = view_rect.width
    v_height = view_rect.height

    words = pagina.get_text("words")
    words_cajetin = []
    for w in words:
        r_nat = fitz.Rect(w[0], w[1], w[2], w[3])
        r_view = r_nat * mat_view
        if r_view.x0 >= v_width * 0.45 and r_view.y0 >= v_height * 0.50:
            words_cajetin.append(w[4])

    texto_roi = " ".join(words_cajetin)
    match_cod = re.search(r'\b[A-Z0-9]{2,}(?:-[A-Z0-9]+){2,}(?:-R\d+|-REV\d+)?\b', texto_roi)
    if match_cod:
        return match_cod.group(0)

    texto_completo = pagina.get_text("text")
    patrones_codigo = [
        r'\b[A-Z0-9]{2,}(?:-[A-Z0-9]+){2,}(?:-R\d+|-REV\d+)?\b',
        r'\b\d+-[A-Z0-9]+-[0-9-]+(?:-R\d+|-REV\d+)?\b'
    ]
    for pat in patrones_codigo:
        matches = re.findall(pat, texto_completo)
        if matches:
            falsos_positivos = ["ESCALA", "FECHA", "PROYECTO", "TUBOS", "INDICADA", "DIBUJO", "PLANO", "REVISION"]
            candidatos = [m.strip() for m in matches if not any(fp in m.upper() for fp in falsos_positivos) and len(m.strip()) > 6]
            if candidatos:
                return candidatos[0]

    return "No detectado"

def extraer_descripcion_plano(pagina, codigo_plano="No detectado"):
    rot = pagina.rotation
    rect_pag = pagina.rect
    if rot in [90, 270]:
        ancho_v, alto_v = rect_pag.height, rect_pag.width
    else:
        ancho_v, alto_v = rect_pag.width, rect_pag.height

    words = pagina.get_text("words")
    mat_directa = pagina.rotation_matrix

    rect_proyecto_vista = None
    for w in words:
        texto_word = w[4].upper().strip()
        if "PROYECTO" in texto_word:
            r_nativo = fitz.Rect(w[0], w[1], w[2], w[3])
            r_vista = r_nativo * mat_directa
            if r_vista.y0 > (alto_v * 0.50):
                rect_proyecto_vista = r_vista
                break

    titulo_plano = "No detectado"

    if rect_proyecto_vista:
        view_box_desc = fitz.Rect(
            rect_proyecto_vista.x0 - 40,
            rect_proyecto_vista.y1 + 2,
            rect_proyecto_vista.x0 + 380,
            rect_proyecto_vista.y1 + 110
        )
        mat_inversa = pagina.derotation_matrix
        nat_box_desc = view_box_desc * mat_inversa

        texto_celda = pagina.get_text("text", clip=nat_box_desc).strip()
        lineas = [l.strip() for l in texto_celda.split('\n') if l.strip()]

        lineas_limpias = []
        for l in lineas:
            l_limp = limpiar_texto_comillas(l)
            l_up = l_limp.upper()
            palabras_descarte = [
                "ESCALA", "FECHA", "PROYECTO", "INDICADA", "CODIGO", "CÓDIGO", 
                "500M", "100M", "200M", "300M", "400M", "ESCALA GRAFICA", "REVISION"
            ]
            if l_limp and not any(k in l_up for k in palabras_descarte) and l_limp != codigo_plano:
                if not re.match(r'^[\+\-]?\d+[\.\,\+\d]*\s*m?$', l_limp) and len(l_limp) > 3:
                    lineas_limpias.append(l_limp)

        if lineas_limpias:
            titulo_plano = " - ".join(lineas_limpias[:3])

    if titulo_plano == "No detectado":
        view_roi_cajetin = fitz.Rect(ancho_v * 0.55, alto_v * 0.70, ancho_v, alto_v)
        nat_roi_cajetin = view_roi_cajetin * pagina.derotation_matrix

        texto_roi = pagina.get_text("text", clip=nat_roi_cajetin)
        lineas_fallback = [l.strip() for l in texto_roi.split('\n') if len(l.strip()) > 3]

        lineas_validas = []
        for l in lineas_fallback:
            l_limp = limpiar_texto_comillas(l)
            l_up = l_limp.upper()
            palabras_descarte = [
                "ESCALA", "FECHA", "PROYECTO", "INDICADA", "CODIGO", "MTC", 
                "MINISTERIO", "INTERSUR", "ESCALA GRAFICA", "500M", "400M", "300M", "200M", "100M"
            ]
            if l_limp and not any(k in l_up for k in palabras_descarte) and l_limp != codigo_plano:
                if not re.match(r'^[\+\-]?\d+[\.\,\+\d]*\s*m?$', l_limp) and len(l_limp) > 3:
                    lineas_validas.append(l_limp)

        if lineas_validas:
            titulo_plano = " - ".join(lineas_validas[:3])

    if titulo_plano != "No detectado":
        titulo_plano = limpiar_texto_comillas(titulo_plano)

    return titulo_plano

def extraer_datos_inteligentes_cajetin(pagina):
    codigo_plano = extraer_codigo_plano(pagina)
    titulo_plano = extraer_descripcion_plano(pagina, codigo_plano)
    return codigo_plano, titulo_plano

def buscar_posicion_espacio_libre(imagen_gris, ancho_sello, alto_sello, sellos_ya_puestos, es_a4, paso=10):
    _, binaria = cv2.threshold(imagen_gris, 230, 255, cv2.THRESH_BINARY_INV)
    alto_img, ancho_img = binaria.shape
    area_sello = ancho_sello * alto_sello
    pad = 20

    candidatos = []

    if es_a4:
        for y in range(alto_img - alto_sello - 30, 30, -paso):
            for x in range(30, ancho_img - ancho_sello - 30, paso):
                if x > (ancho_img * 0.60) and y > (alto_img * 0.70):
                    continue

                colision = False
                for (sx, sy, sw, sh) in sellos_ya_puestos:
                    if not (x + ancho_sello + pad < sx or x > sx + sw + pad or
                            y + alto_sello + pad < sy or y > sy + sh + pad):
                        colision = True
                        break
                if colision:
                    continue

                caja = binaria[y : y + alto_sello, x : x + ancho_sello]
                pixeles_ocupados = cv2.countNonZero(caja)
                porcentaje_libre = 1.0 - (pixeles_ocupados / float(area_sello))

                score_libre = porcentaje_libre * 100.0
                factor_x = 1.0 - (x / float(ancho_img))
                factor_y = y / float(alto_img)
                bonif_posicion = (factor_x * 10.0) + (factor_y * 10.0)

                score_total = score_libre + bonif_posicion
                candidatos.append((score_total, porcentaje_libre, x, y))

        if not candidatos:
            return 30, alto_img - alto_sello - 30
    else:
        for x in range(ancho_img - ancho_sello - 30, 30, -paso):
            for y in range(alto_img - alto_sello - 30, 30, -paso):
                if x > (ancho_img * 0.60) and y > (alto_img * 0.70):
                    continue

                colision = False
                for (sx, sy, sw, sh) in sellos_ya_puestos:
                    if not (x + ancho_sello + pad < sx or x > sx + sw + pad or
                            y + alto_sello + pad < sy or y > sy + sh + pad):
                        colision = True
                        break
                if colision:
                    continue

                caja = binaria[y : y + alto_sello, x : x + ancho_sello]
                pixeles_ocupados = cv2.countNonZero(caja)
                porcentaje_libre = 1.0 - (pixeles_ocupados / float(area_sello))

                score_libre = porcentaje_libre * 100.0
                factor_x = x / float(ancho_img)
                factor_y = y / float(alto_img)
                bonif_posicion = (factor_x * 5.0) + (factor_y * 5.0)

                score_total = score_libre + bonif_posicion
                candidatos.append((score_total, porcentaje_libre, x, y))

        if not candidatos:
            return 50, 50

    candidatos.sort(key=lambda item: item[0], reverse=True)
    mejor_score, mejor_libre, mejor_x, mejor_y = candidatos[0]

    return mejor_x, mejor_y

def agregar_sello_png_dinamico(pagina, view_rect, rect_nativo, nombre_png, color_rgb, texto_fecha, rotacion):
    ruta_png = None
    posibles_rutas = [nombre_png, os.path.join(CARPETA_SELLOS, nombre_png)]

    for r in posibles_rutas:
        if os.path.exists(r):
            ruta_png = r
            break

    if not ruta_png:
        return False, f"No se encontró la imagen '{nombre_png}'."

    pagina.insert_image(rect_nativo, filename=ruta_png, rotate=rotacion)

    partes = re.split(r'[/.-]', texto_fecha.strip())
    if len(partes) >= 3:
        dia_txt, mes_txt, anio_txt = partes[0].zfill(2), partes[1].zfill(2), partes[2][-2:].zfill(2)
    else:
        dia_txt, mes_txt, anio_txt = "", "", ""

    fontname = "hebo"
    fontsize = 0.50 * CM_TO_PT

    y_vista = view_rect.y0 + (view_rect.height * 0.81)
    posiciones = [(dia_txt, 0.38), (mes_txt, 0.62), (anio_txt, 0.89)]

    for txt, rel_x in posiciones:
        if not txt:
            continue
        ancho_txt = fitz.get_text_length(txt, fontname=fontname, fontsize=fontsize)
        x_vista = view_rect.x0 + (view_rect.width * rel_x) - (ancho_txt / 2.0)

        pt_vista = fitz.Point(x_vista, y_vista)
        pt_nativo = pt_vista * pagina.derotation_matrix

        pagina.insert_text(pt_nativo, txt, fontsize=fontsize, fontname=fontname, color=color_rgb, rotate=rotacion)
        pt_nativo_offset = fitz.Point(pt_nativo.x + 0.3, pt_nativo.y)
        pagina.insert_text(pt_nativo_offset, txt, fontsize=fontsize, fontname=fontname, color=color_rgb, rotate=rotacion)

    return True, None

def procesar_pdf(pdf_bytes, lista_sellos_elegidos, libreria_archivos, texto_fecha, paso=10):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
        tmp_pdf.write(pdf_bytes)
        path_pdf = tmp_pdf.name

    doc = fitz.open(path_pdf)
    resumen_planos = []
    alertas_sellos = []
    contiene_a4 = False

    ancho_sello_pt = ANCHO_SELLO_CM * CM_TO_PT
    alto_sello_pt = ALTO_SELLO_CM * CM_TO_PT

    for i in range(len(doc)):
        pagina = doc[i]
        rot = pagina.rotation

        cod_plano, titulo_plano = extraer_datos_inteligentes_cajetin(pagina)
        resumen_planos.append({
            "Hoja": i + 1,
            "Código de Plano": cod_plano,
            "Título / Descripción": titulo_plano
        })

        pixmap = pagina.get_pixmap(dpi=100)
        img_np = np.frombuffer(pixmap.samples, dtype=np.uint8).reshape(pixmap.h, pixmap.w, pixmap.n)
        imagen_gris = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY) if pixmap.n >= 3 else img_np

        alto_img, ancho_img = imagen_gris.shape
        max_dim = max(ancho_img, alto_img)
        
        es_hoja_a4 = max_dim < 1300
        if es_hoja_a4:
            contiene_a4 = True

        ancho_sello_px = int(ANCHO_SELLO_CM * 39.37007874)
        alto_sello_px = int(ALTO_SELLO_CM * 39.37007874)
        sellos_puestos_hoja = []

        for item_sello in lista_sellos_elegidos:
            x_px, y_px = buscar_posicion_espacio_libre(
                imagen_gris, ancho_sello_px, alto_sello_px, sellos_puestos_hoja, es_a4=es_hoja_a4, paso=paso
            )

            x0_pt_view = x_px * 0.72
            y0_pt_view = y_px * 0.72

            view_rect = fitz.Rect(x0_pt_view, y0_pt_view, x0_pt_view + ancho_sello_pt, y0_pt_view + alto_sello_pt)
            rect_nativo = view_rect * pagina.derotation_matrix

            exito = True
            msg_err = None

            if item_sello == "CC - Copia Controlada (Rojo)":
                exito, msg_err = agregar_sello_png_dinamico(
                    pagina, view_rect, rect_nativo, "cc_sin_fondo.png", (0.0, 0.20, 0.65), texto_fecha, rot
                )
            elif item_sello == "CI - Copia Informativa (Azul)":
                exito, msg_err = agregar_sello_png_dinamico(
                    pagina, view_rect, rect_nativo, "ci_sin_fondo.png", (0.80, 0.0, 0.0), texto_fecha, rot
                )
            else:
                ruta_img = libreria_archivos.get(item_sello)
                if ruta_img and os.path.exists(ruta_img):
                    pagina.insert_image(rect_nativo, filename=ruta_img, rotate=rot)
                else:
                    exito = False
                    msg_err = f"No se encontró la imagen '{item_sello}'."

            if exito:
                sellos_puestos_hoja.append((x_px, y_px, ancho_sello_px, alto_sello_px))
            else:
                if msg_err and msg_err not in alertas_sellos:
                    alertas_sellos.append(msg_err)

    output_pdf_path = path_pdf.replace(".pdf", "_SELLADO.pdf")
    doc.save(output_pdf_path)
    doc.close()

    with open(output_pdf_path, "rb") as f:
        pdf_final_bytes = f.read()

    os.remove(path_pdf)
    os.remove(output_pdf_path)

    return pdf_final_bytes, resumen_planos, alertas_sellos, contiene_a4

# --- PRINCIPAL ---
def main():
    st.set_page_config(page_title="AO RNLD US - Estampador & GR", page_icon="📐", layout="wide")
    inicializar_estado_sesion()

    if not st.session_state.autenticado:
        vista_login()
        return

    aplicar_estilos_custom()

    # --- BARRA LATERAL ---
    if os.path.exists(RUTA_LOGO):
        st.sidebar.image(RUTA_LOGO, use_container_width=True)
    st.sidebar.markdown(f"**Usuario:** `{st.session_state.usuario}`")
    if st.sidebar.button("Cerrar Sesión", use_container_width=True):
        st.session_state.autenticado = False
        st.session_state.usuario = None
        st.rerun()

    st.sidebar.divider()
    st.sidebar.header("Ajustes de Escaneo")
    paso_evaluacion = st.sidebar.slider("Precisión (Paso en px):", 5, 30, 10, 5)

    st.sidebar.header("Cargar Nuevas Firmas")
    with st.sidebar.expander("Subir Imagen a Base de Datos", expanded=False):
        nuevo_nombre = st.sidebar.text_input("Nombre de la firma/sello:")
        archivo_nuevo = st.sidebar.file_uploader("Subir imagen (PNG/JPG):", type=["png", "jpg", "jpeg"])

        if st.sidebar.button("Guardar Sello", use_container_width=True):
            if nuevo_nombre.strip() and archivo_nuevo:
                ext = archivo_nuevo.name.split(".")[-1]
                ruta_dest = os.path.join(CARPETA_SELLOS, f"{nuevo_nombre.lower().strip().replace(' ', '_')}.{ext}")
                with open(ruta_dest, "wb") as f:
                    f.write(archivo_nuevo.read())
                st.sidebar.success("¡Sello Guardado!")
                st.rerun()

    # --- INTERFAZ PRINCIPAL ---
    st.title("SISTEMA DE ESTAMPADO Y GUÍAS DE REMISIÓN")
    st.caption("AO RNLD US • INGENIERÍA • DISEÑO • CONSTRUCCIÓN")

    col1, col2, col3 = st.columns([1.2, 1, 1])

    with col1:
        archivo_pdf = st.file_uploader("1. Selecciona tu PDF consolidado:", type=["pdf"])
        secuencia_gr = st.text_input("2. Secuencia Base (ej: 8418-OSP-SG-2026):", value="8418-OSP-SG-2026")

    with col2:
        generar_guias_opcion = st.checkbox("Generar Guías de Remisión (Solo A3)", value=True)
        fecha_obj = st.date_input("3. Fecha de Sellado (J5):", value=datetime.date.today(), format="DD/MM/YYYY")
        texto_fecha = fecha_obj.strftime("%d/%m/%Y")

    libreria_archivos = obtener_libreria_sellos()
    opciones_sellos = ["CC - Copia Controlada (Rojo)", "CI - Copia Informativa (Azul)"] + list(libreria_archivos.keys())

    with col3:
        areas_opciones = ["SUPERVISION", "CALIDAD", "TOPOGRAFIA", "PRODUCCION"]
        areas_seleccionadas = st.multiselect(
            "4. Áreas para Guías:",
            options=areas_opciones,
            default=areas_opciones,
            disabled=not generar_guias_opcion
        )

        sellos_seleccionados = st.multiselect(
            "5. Selecciona Sellos/Firmas:",
            options=opciones_sellos,
            default=[opciones_sellos[0]]
        )

    st.divider()

    if archivo_pdf and sellos_seleccionados:
        if st.button("Estampar PDF y Procesar Documentos", use_container_width=True):
            with st.spinner("Procesando plano con algoritmos adaptativos..."):
                pdf_res, resumen, alertas_sellos, es_a4 = procesar_pdf(
                    archivo_pdf.read(), 
                    sellos_seleccionados, 
                    libreria_archivos, 
                    texto_fecha,
                    paso=paso_evaluacion
                )

                rev_tag = "REV"
                if resumen and resumen[0]["Código de Plano"] != "No detectado":
                    rev_num = extraer_numero_revision(resumen[0]["Código de Plano"])
                    if rev_num:
                        rev_tag = f"R{rev_num}"

                excels_generados = {}
                lista_pdfs_guias = []
                errores_pdf = []

                if generar_guias_opcion and not es_a4:
                    for idx, area in enumerate(areas_seleccionadas):
                        excel_bytes, secuencia_inc = generar_excel_por_area(
                            resumen_planos=resumen,
                            fecha_texto=texto_fecha,
                            secuencia_base=secuencia_gr,
                            area=area,
                            offset_correlativo=idx
                        )
                        if excel_bytes:
                            nombre_excel = f"{secuencia_inc}_{rev_tag}_{area}.xlsx"
                            pdf_excel, error_msg = convertir_excel_a_pdf(excel_bytes)

                            if pdf_excel:
                                lista_pdfs_guias.append(pdf_excel)
                            elif error_msg:
                                errores_pdf.append(f"{area}: {error_msg}")

                            excels_generados[area] = {
                                "bytes": excel_bytes,
                                "nombre": nombre_excel,
                                "secuencia": secuencia_inc,
                                "pdf_bytes": pdf_excel
                            }

                st.session_state['pdf_res'] = pdf_res
                st.session_state['resumen'] = resumen
                st.session_state['excels_generados'] = excels_generados
                st.session_state['pdf_nombre'] = f"{secuencia_gr}_{rev_tag}_PLANOS_SELLADOS.pdf"
                st.session_state['errores_pdf'] = errores_pdf
                st.session_state['alertas_sellos'] = alertas_sellos
                st.session_state['es_a4'] = es_a4
                st.session_state['guias_activadas'] = generar_guias_opcion

                if lista_pdfs_guias:
                    st.session_state['pdf_guias_unificado'] = unificar_pdfs(lista_pdfs_guias)
                else:
                    st.session_state['pdf_guias_unificado'] = None

    if 'resumen' in st.session_state:
        st.success("¡Planos estampados y procesados con éxito!")

        if st.session_state.get('es_a4'):
            st.info("Se detectó formato A4: La búsqueda se ejecutó en 'L Invertida' y no se generaron Guías de Remisión.")
        elif not st.session_state.get('guias_activadas'):
            st.info("La opción de Guías estuvo desactivada.")

        if st.session_state.get('alertas_sellos'):
            for alert in st.session_state['alertas_sellos']:
                st.warning(alert)

        st.subheader("Descargas Generales")
        col_dl1, col_dl2 = st.columns(2)

        with col_dl1:
            st.download_button(
                "Descargar PDF Planos Sellados", 
                data=st.session_state['pdf_res'], 
                file_name=st.session_state['pdf_nombre'], 
                mime="application/pdf", 
                use_container_width=True
            )

        with col_dl2:
            if st.session_state.get('pdf_guias_unificado'):
                st.download_button(
                    "Descargar PDF Consolidado de Guías", 
                    data=st.session_state['pdf_guias_unificado'], 
                    file_name=f"GUIAS_REMISION_CONSOLIDADAS_{datetime.datetime.now().strftime('%Y%m%d')}.pdf", 
                    mime="application/pdf", 
                    use_container_width=True
                )

        if st.session_state.get('excels_generados'):
            st.divider()
            st.markdown("#### Archivos Excel y PDF de Guías por Área:")
            cols_excels = st.columns(len(st.session_state['excels_generados']))

            for idx, (area, data_excel) in enumerate(st.session_state['excels_generados'].items()):
                with cols_excels[idx]:
                    st.download_button(
                        label=f"Excel: {data_excel['secuencia']} ({area})",
                        data=data_excel["bytes"],
                        file_name=data_excel["nombre"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"excel_{area}"
                    )
                    if data_excel.get("pdf_bytes"):
                        st.download_button(
                            label=f"PDF: {data_excel['secuencia']} ({area})",
                            data=data_excel["pdf_bytes"],
                            file_name=f"{data_excel['secuencia']}_{area}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"pdf_{area}"
                        )

        st.divider()
        st.subheader("Resumen de Planos Detectados")
        st.dataframe(st.session_state['resumen'], use_container_width=True)

if __name__ == "__main__":
    main()